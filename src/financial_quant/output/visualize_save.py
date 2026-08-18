import os
import sys
import re
import pandas as pd
import numpy as np
from pathvalidate import sanitize_filepath
from matplotlib import pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
from IPython.display import display, Markdown as md
from pathlib import Path
import altair as alt

def one_y_axis(x_data, y_data_list, title="", xlabel="", ylabel="", 
                     series_labels=None, markers=None, colors=None,
                     figure_size=(10, 6), y_limits=None, 
                     save_config=None, fill_config=None):
    """
    Plots data on a single y-axis.

    Args:
        x_data (array-like): Data for the x-axis.
        y_data_list (list of array-like): A list of datasets for the y-axis.
        title (str): The title of the graph.
        xlabel (str): The label for the x-axis.
        ylabel (str): The label for the y-axis.
        series_labels (list of str, optional): Identifiers for each data series. 
        markers (list of str, optional): The markers to use for each series.
        colors (list of str, optional): Colors for each series.
        figure_size (tuple): The width and height of the figure in inches.
        y_limits (tuple, optional): The (min, max) values for the y-axis.
        save_config (dict, optional): Config for saving the file. Keys: 'volume', 'chapter', 'file_name'.
        fill_config (dict, optional): Config for filling areas. Keys: 'Between' (list of 1 or 2 indices),
                                      'Start', 'End', 'Colors', 'Labels', 'Alpha'.
    """


    num_series = len(y_data_list)

    # --- Smart Defaults (Frictionless Inputs) ---
    series_labels = series_labels or [f"Series {i+1}" for i in range(num_series)]
    markers = markers or [""] * num_series
    colors = colors or plt.cm.viridis_r(np.linspace(0, 1, num_series))

    # Input Validation
    if not (len(series_labels) == len(markers) == len(colors) == num_series):
        raise ValueError("Lengths of 'series_labels', 'markers', and 'colors' must match 'y_data_list'.")

    # --- Plotting Setup (Protects Global State) ---
    with plt.style.context('ggplot'):
        fig, ax = plt.subplots(figsize=figure_size)
        fig.suptitle(title)

        # --- Pythonic Loop ---
        for y_data, label, marker, color in zip(y_data_list, series_labels, markers, colors):
            ax.plot(x_data, y_data, label=label, marker=marker, color=color)

        # --- Implement the Missing Fill Feature ---
        if fill_config:
            indices = fill_config.get('Between', [0])
            y1 = y_data_list[indices[0]]
            y2 = y_data_list[indices[1]] if len(indices) > 1 else np.zeros_like(y1)
            
            start, end = fill_config.get('Start', 0), fill_config.get('End', len(x_data))
            
            ax.fill_between(
                x_data[start:end], y1[start:end], y2[start:end],
                color=fill_config.get('Colors', 'gray'),
                alpha=fill_config.get('Alpha', 0.3),
                label=fill_config.get('Labels', None)
            )

        # --- Final Touches ---
        if y_limits:
            ax.set_ylim(y_limits)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.legend()
        plt.tight_layout()

        # --- Save Figure ---
        if save_config:
            # Assuming save_results is imported/defined elsewhere in your script
            path = save_results(save_config=save_config)
            if path:
                plt.savefig(path, dpi=300, bbox_inches='tight')

        plt.show()
def create_workbook(df,sheet_name='sheet1', save_config=None):
    """
    Writes a DataFrame to a specific sheet in an Excel workbook and auto-fits
    column widths for readability.


    Args:
        sheet_name (str): The name of the sheet to create or replace.
        df (pd.DataFrame): The DataFrame to write.
        save_config (dict, optional): Configuration for saving the file, passed
         Keys: 'volume':folder, 'chapter':'subfolder, 'file_name':file name. Defaults to {}.     
    """



    
    # Fix mutable default argument
    if save_config is None:
        save_config = {}


    # --- 1. Sanitize Sheet Name ---
    sane_sheet_name = re.sub(r'[\\*?:/\[\]]', '', str(sheet_name))
    if len(sane_sheet_name) > 31:
        sane_sheet_name = sane_sheet_name[:31]
    if not sane_sheet_name:
        sane_sheet_name = "Sheet1"


    # --- 2. Get Save Path ---
    file_name = save_config.get('file_name', 'output.xlsx')
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'
    save_config['file_name']=file_name
    # Assuming save_results is defined elsewhere in your code
    try:
        path_filename = save_results(save_config=save_config)
        if path_filename is None:
            path_filename = file_name
    except NameError:
        # Fallback if save_results is not defined
        path_filename = file_name


    # --- 3. Write and Format ---
    try:
        # Create empty workbook if it doesn't exist to allow 'append' mode
        if not os.path.exists(path_filename):
            pd.DataFrame().to_excel(path_filename, sheet_name="Sheet1")


        # Write DataFrame
        with pd.ExcelWriter(
            path_filename,
            mode='a',
            engine='openpyxl',
            if_sheet_exists='replace',
            datetime_format='YYYY-MM-DD'
        ) as writer:
            # FIXED: Using sane_sheet_name instead of sheet_name
            df.to_excel(writer, sheet_name=sane_sheet_name, index=True)


        # Format with openpyxl
        workbook = openpyxl.load_workbook(path_filename)
        
        # FIXED: Using sane_sheet_name
        try:
            ws = workbook[sane_sheet_name]
        except KeyError:
            print(f"Error: Sheet '{sane_sheet_name}' not found after writing.")
            return


        # Auto-fit columns
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)


            for cell in column_cells:
                try:
                    if cell.value: # Check if cell is not empty
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass


            ws.column_dimensions[column_letter].width = max_length + 2


        # Delete default Sheet1 if necessary
        if 'Sheet1' in workbook.sheetnames and sane_sheet_name != 'Sheet1' and len(workbook.sheetnames) > 1:
            del workbook['Sheet1']
            
        workbook.save(path_filename)
        display(md(f"### ***✅ Successfully wrote and formatted sheet {sane_sheet_name} in {path_filename}***"))
        
    except Exception as e:
        display(md("### ❌ **ERROR during Excel write/format:**"))
        print(f"Exception details: {e}")

def save_results(save_config: dict = None):
    """
    Interactively prompts the user to confirm and generate a safe save path.


    This function pauses execution and asks the user (y/n) if they want to save
    a file. It is environment-aware:


    - **In Google Colab:** It attempts to use '/content/drive/MyDrive'. If not
      mounted, it will try to mount it. If mounting fails, it falls back
      to the temporary '/content' directory and issues a warning.
    - **In a local environment:** It uses the current working directory.


    The function constructs a full path from the base folder and the optional
    'volume' and 'chapter' subdirectories. All path components are sanitized
    using `pathvalidate`.


    Args:
        save_config (dict, optional): A dictionary containing path components.
            'volume' (str, optional): The name of a top-level subdirectory.
            'chapter' (str, optional): The name of a nested subdirectory.
            'file_name' (str, optional): The final file name.
                Defaults to 'output.txt' if not provided or if sanitization
                results in an empty string.


    Returns:
        str or None: A complete, sanitized, absolute string path to the
            file if the user confirms 'y' and path creation succeeds.
            Returns None if the user chooses 'n' or if an error occurs.
    
    Example:
        config = {
            'volume': 'My_Notebooks',
            'chapter': 'Chapter_01',
            'file_name': 'results.json'
        }
        save_path = save_results(config)
        
        # User inputs 'y'
        
        if save_path:
            # save_path might be '/content/drive/MyDrive/My_Notebooks/Chapter_01/results.json'
            print(f"Saving to: {save_path}")
            # ... proceed to write file ...
    """



    save_config = save_config or {}


    # --- Handle IPython imports safely ---
    is_colab = 'google.colab' in sys.modules
    try:
        from IPython.display import display, Markdown as md
        def display_msg(text): display(md(text))
    except ImportError:
        def display_msg(text): print(text.replace('###', '').replace('*', ''))


    # --- 1. Get user's choice (y/n) ---
    default_value = "n"
    prompt = "❓ Do you want to save the file? (y/n) (press enter for n): "


    try:
        raw_input = input(prompt).strip().lower()
        choice = raw_input if raw_input else default_value
    except (EOFError, KeyboardInterrupt):
        return None


    while choice not in ['y', 'n']:
        choice = input(prompt).strip().lower()


    # --- 2. Handle "No" ---
    if choice == 'n':
        display_msg('### ***❌ File Not Saved***.')
        return None


    # --- 3. Handle "Yes" ---
    display_msg("### ***⌛ Generating A Path***")


    # --- Get and sanitize config values ---
    volume = sanitize_filepath(save_config.get('volume', ''))
    chapter = sanitize_filepath(save_config.get('chapter', ''))
    
    # The caller should provide the extension, but we fall back to a generic name just in case
    file_name = sanitize_filepath(save_config.get('file_name', 'output'))
    if not file_name:
        file_name = 'output'


    subfolder = os.path.join(volume, chapter)


    # --- Drive/Folder Logic ---
    base_folder = ''
    drive_path = '/content/drive/MyDrive'
    
    if is_colab:
        if os.path.exists(drive_path):
            base_folder = drive_path
        else:
            try:
                from google.colab import drive
                drive.mount('/content/drive')
                base_folder = drive_path
            except Exception as e:
                base_folder = '/content'
                display_msg(f"### ⚠️ **Drive Mount Failed:** {e}. Saving to temporary '/content' folder.")
    else:
        base_folder = os.getcwd()


    # --- Path Creation ---
    try:
        full_folder_path = os.path.join(base_folder, subfolder)
        path_obj = Path(full_folder_path)
        path_obj.mkdir(parents=True, exist_ok=True)
        
        final_path_str = os.path.join(str(path_obj), file_name)


        display_msg(f'### ✅ **File Path Generated:**\n`{final_path_str}`')
        
        if base_folder == '/content':
            display_msg('### ⚠️ *File is in a temporary location and will be lost on runtime restart.*')


        return final_path_str


    except Exception as e:
        display_msg('### ❌ **ERROR Creating Directory:**')
        print(e)
        return None


def graph_uncertainty_arb(df_list, asset):
    """
    Combines a list of DataFrames and generates side-by-side Altair
    line charts for Trading Volume and Closing Price.
    """
    # 1. Prepare the combined dataset
    df_combined = pd.concat(df_list).reset_index(names='date')

    # Calculate dynamic y-axis domains (adding a 10% buffer)
    price_domain = [df_combined['close'].min() * 0.9, df_combined['close'].max() * 1.1]
    volume_domain = [df_combined['volume'].min() * 0.9, df_combined['volume'].max() * 1.1]

    # 2. Create a base chart with shared properties
    # Using lines for volume avoids messy overlapping bars for multiple symbols
    base_chart = alt.Chart(df_combined).mark_line().encode(
        x=alt.X('date:T', title='Date'),
        color=alt.Color('symbol:N', legend=alt.Legend(title="Symbol"))
    )

    # 3. Generate the Price Chart
    price_chart = base_chart.encode(
        y=alt.Y('close:Q', title='Price', scale=alt.Scale(domain=price_domain))
    ).properties(
        width=300,
        height=350,
        title=f'{asset} Closing Prices - Arb Effect'
    )

    # 4. Generate the Volume Chart
    volume_chart = base_chart.encode(
        y=alt.Y('volume:Q', title='Volume', scale=alt.Scale(domain=volume_domain))
    ).properties(
        width=300,
        height=350,
        title=f'{asset} Trading Volume - Uncertainty Transmission'
    )

    # 5. Combine side-by-side using the | operator
    side_by_side_chart = volume_chart | price_chart

    # Returning the chart is best practice in Jupyter environments
    # It allows you to save the output to a variable when calling the function
    return side_by_side_chart


